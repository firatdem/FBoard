# json_to_excel.py
import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook import Workbook


def convert_json_to_excel(json_file="output.json", excel_file="output.xlsx"):
    """
    Converts JSON data (with 'employees' and 'job_sites' keys) into an Excel workbook.
    It creates three sheets:
      1. "Job Site Summary" – a summary of electrician counts per job site with styling and hyperlinks.
      2. "Employees" – a sheet containing all employee data.
      3. "Employee List" – a grouped list of employees by job site.

    Parameters:
        json_file (str): Path to the input JSON file.
        excel_file (str): Path for the output Excel file.
    """
    # Load JSON data from file
    with open(json_file, "r") as file:
        data = json.load(file)

    # Create DataFrames for employees and job sites
    employees = pd.DataFrame(data["employees"])
    job_sites = pd.DataFrame(data["job_sites"])

    # Initialize a list for organized job site data
    job_site_details = []

    # Process each job site
    for _, job_site in job_sites.iterrows():
        site_name = job_site["name"]
        # Exclude employees whose current_status is "Sick"
        site_employees = employees[
            (employees["job_site"] == site_name) & (employees["current_status"] != "Sick")
            ]

        # Filter electricians (case-insensitive search for "Electrician")
        electrician_roles = site_employees[
            site_employees["role"].str.contains("Electrician", na=False, case=False)
        ]

        # Total number of electricians
        total_electricians = electrician_roles.shape[0]

        # Count electricians by skill by exploding the 'skills' list
        skill_counts = (
            electrician_roles.explode("skills")["skills"]
            .value_counts()
            .to_dict()
        )

        # Build a dictionary of job site data with electrician counts per skill
        job_site_data = {"Job Site": site_name}
        for skill, count in skill_counts.items():
            job_site_data[f"Electricians ({skill})"] = count
        job_site_data["Total Electricians"] = total_electricians

        job_site_details.append(job_site_data)

    # Convert job site details into a DataFrame
    job_site_summary = pd.DataFrame(job_site_details)

    # If "Total Electricians" column is missing, calculate it from other columns
    if "Total Electricians" not in job_site_summary.columns:
        job_site_summary["Total Electricians"] = job_site_summary.apply(
            lambda row: sum(
                [row[col] for col in job_site_summary.columns
                 if col.startswith("Electricians") and not pd.isna(row[col])]
            ),
            axis=1
        )

    # Define a styler for the "Total Electricians" column (light yellow background, bold text)
    styler = job_site_summary.style.set_properties(
        subset=["Total Electricians"],
        **{"font-weight": "bold", "background-color": "#FFFFE0"}
    )

    # Export to Excel with styling and hyperlinks using OpenPyXL
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        # Write Job Site Summary sheet with styling
        styler.to_excel(writer, sheet_name="Job Site Summary", index=False)
        # Write Employees sheet (all employee data)
        employees.to_excel(writer, sheet_name="Employees", index=False)

        # Create the Employee List sheet
        workbook = writer.book
        employee_list_sheet = workbook.create_sheet("Employee List")

        # Dictionary to track row numbers per job site for hyperlink targets
        job_site_row_mapping = {}
        current_row = 1

        # Write headers for the Employee List sheet
        headers = ["Job Site", "Employee Name", "Role", "Skills"]
        for col_num, header in enumerate(headers, 1):
            cell = employee_list_sheet.cell(row=current_row, column=col_num, value=header)
            cell.font = Font(bold=True)
        current_row += 1

        # Group employees by job site and write their details
        for job_site_name, group in employees.groupby("job_site"):
            # Record starting row for this job site group
            job_site_row_mapping[job_site_name] = current_row

            # Write job site name as a header in bold
            employee_list_sheet.cell(row=current_row, column=1, value=job_site_name).font = Font(bold=True)
            current_row += 1

            # Write each employee's details
            for _, emp in group.iterrows():
                employee_list_sheet.cell(row=current_row, column=2, value=emp["text"])
                employee_list_sheet.cell(row=current_row, column=3, value=emp["role"])
                skills = ", ".join(emp["skills"]) if isinstance(emp["skills"], list) else ""
                employee_list_sheet.cell(row=current_row, column=4, value=skills)
                current_row += 1
            current_row += 1  # Add an empty row after each group

        # Add hyperlinks in the Job Site Summary sheet linking to the Employee List sheet
        summary_sheet = workbook["Job Site Summary"]
        for row in range(2, summary_sheet.max_row + 1):
            cell = summary_sheet.cell(row=row, column=1)  # Assuming Job Site names are in column A
            job_site_name = cell.value
            if job_site_name in job_site_row_mapping:
                target_row = job_site_row_mapping[job_site_name]
                cell.hyperlink = f"#'Employee List'!A{target_row}"
                cell.style = "Hyperlink"

        # Rearrange the sheet order to: Job Site Summary, Employee List, Employees
        sheet_order = ["Job Site Summary", "Employee List", "Employees"]
        for idx, sheet_name in enumerate(sheet_order):
            workbook.move_sheet(sheet_name, offset=idx)

    print(job_site_summary)
    print(f"\nJSON data has been converted to Excel and saved as {excel_file}")


# To run as a standalone script:
if __name__ == "__main__":
    convert_json_to_excel()
